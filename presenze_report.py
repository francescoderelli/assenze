#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Presenze Base -> Report ore annue + percentuali presenza (ore lavorabili) + didascalia.

Output Excel:
- percentuali presenza (foglio attivo)
- Ore_Annue_Pivot
- Dettaglio_Annuale
- Liste
"""

from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================
# DIZIONARIO CAUSALI
# Categorie: presenza | cig | festivita | ferie | assenza_varie | infortunio | escludi
# =========================
CAUSALE_CATEGORY = {
    # --- CIG
    "CIG CALDO": "cig",
    "CIG CALDO 1 GIORNO": "cig",
    "CIG GELO": "cig",
    "CIG GELO 1 GIORNO": "cig",
    "CIG PIOGGIA": "cig",
    "CIG PIOGGIA 1 GIORNO": "cig",
    "CIG PIOGGIA CON VITTO": "cig",
    "CIG VENTO": "cig",
    "CIG VENTO 1 GIORNO": "cig",
    "CIG VENTO CON VITTO": "cig",
    "LLUVIA": "cig",
    "INTEMPERIE": "cig",

    # --- Festività (NEUTRE)
    "FESTIVITA": "festivita",
    "FESTIVITA 4 NOVEMBRE": "festivita",
    "FESTIVITA DOMENICA": "festivita",

    # --- Ferie
    "FERIE": "ferie",

    # --- Infortunio
    "INFORTUNIO": "infortunio",

    # --- Assenze varie
    "ALLATTAMENTO": "assenza_varie",
    "ASPETTATIVA": "assenza_varie",
    "ASSENZA INGIUSTIFICATA": "assenza_varie",
    "ASSENZA NON RETRIBUITA": "assenza_varie",
    "ASSEMBLEA SINDACALE": "assenza_varie",
    "CONGEDO MATRIMONIALE": "assenza_varie",
    "CONGEDO PARENTALE": "assenza_varie",
    "CONGEDO PATERNITA": "assenza_varie",
    "DONAZIONE SANGUE": "assenza_varie",
    "PATERNITA": "assenza_varie",
    "LUTTO": "assenza_varie",
    "MALATTIA": "assenza_varie",
    "PERMESSO 104": "assenza_varie",
    "PERMESSO NON RETRIBUITO": "assenza_varie",
    "PERMESSO RETRIBUITO": "assenza_varie",
    "PERMESSO SEA": "assenza_varie",
    "PERMISO JUSTIFICADO/RETRIBUIDO": "assenza_varie",
    "SOSPENSIONE CAUTELATIVA": "assenza_varie",
    "SOSPENSIONE DISCIPLINARE": "assenza_varie",
    "CASSA INTEGRAZIONE": "assenza_varie",

    # --- Presenza
    "CORSI": "presenza",
    "FORMAZIONE": "presenza",
    "GIORNATA DI RIPOSO": "presenza",
    "GUASTO MEZZO": "presenza",
    "LAVORO FESTIVO": "presenza",
    "LAVORO NOTTURNO": "presenza",
    "ORDINARIO": "presenza",
    "ORDINARIO GARANZIA": "presenza",
    "ORDINARIO SABATO": "presenza",
    "ORD. SABATO TRASFERTA": "presenza",
    "ORD. SABATO TRASFERTA CON VITTO": "presenza",
    "PER PREVENTIVO": "presenza",
    "RIPOSO COMPENSATIVO": "presenza",
    "SOSPENSIONE PER INIDONEITA": "presenza",
    "STRAORDINARIO": "presenza",
    "STRAORDINARIO FESTIVO": "presenza",
    "STRAORDINARIO GARANZIA": "presenza",
    "STRAORDINARIO IN TRASFERTA": "presenza",
    "STRAORDINARIO NOTTURNO": "presenza",
    "STRAORDINARIO NOTTURNO FESTIVO": "presenza",
    "TRASF CON VITTO": "presenza",
    "TRASFERTA": "presenza",
    "VIAGGIO": "presenza",
    "VISITA MEDICA": "presenza",
    "LAVORO STRAORDINARIO FESTIVO": "presenza",
    "HEURE SUPPL. > 43": "presenza",

    # --- Distacco
    "DISTACCO": "presenza",
    "DISTACCO ORD. SABATO": "presenza",
    "DISTACCO CON VITTO": "presenza",
    "DISTACCO CON VITTO ORD. SABATO": "presenza",
    "DISTACCO CON VITTO STRAORDINARIO": "presenza",
    "DISTACCO IN TRASFERTA": "presenza",
    "DISTACCO IN TRASFERTA ORD SABATO": "presenza",
    "DISTACCO IN TRASFERTA STRAORDINARIO": "presenza",
    "DISTACCO PER SUBAPPALTO": "presenza",
    "DISTACCO STRAORDINARIO": "presenza",
    "ENERGY DISTACCO CON VITTO": "presenza",
}

VALID_CATS = {"presenza", "cig", "festivita", "ferie", "assenza_varie", "infortunio", "escludi"}
BAD_ROWS = {"TOTAL", "TOTALE", "GRAND TOTAL", "TOTALS", "TOTALE GENERALE"}


def categorize_with_rules(cu: str) -> str | None:
    """Regole robuste per varianti/lingue/punteggiatura."""
    if cu.startswith("CIG"):
        return "cig"
    if ("LLUVIA" in cu) or ("INTEMPERIE" in cu):
        return "cig"
    if cu.startswith("FESTIVIT"):
        return "festivita"
    if "HEURE SUPPL" in cu:
        return "presenza"
    if ("PERMISO" in cu) and ("RETRIBUIDO" in cu):
        return "assenza_varie"
    if "DISTACCO" in cu:
        return "presenza"
    if "PATERNIT" in cu:
        return "assenza_varie"
    if "FORMAZION" in cu:
        return "presenza"
    return None


def find_header_row(raw: pd.DataFrame, max_rows: int = 200) -> int:
    for r in range(min(max_rows, raw.shape[0])):
        row = raw.iloc[r].astype(str).str.strip().str.upper().tolist()
        if "RISORSA" in row and "CAUSALE" in row:
            return r
    raise ValueError("Non trovo l’intestazione con 'Risorsa' e 'Causale' nelle prime righe.")


def find_cols(raw: pd.DataFrame, header_row: int) -> tuple[int, int, int]:
    row = raw.iloc[header_row].astype(str).str.strip().str.upper().tolist()
    ris_col = row.index("RISORSA")
    cau_col = row.index("CAUSALE")

    total_col = None
    for r in range(max(0, header_row - 20), header_row + 1):
        rr = raw.iloc[r].astype(str).str.strip().str.upper().tolist()
        if "TOTAL" in rr:
            total_col = rr.index("TOTAL")

    if total_col is None:
        total_col = raw.shape[1] - 1

    return ris_col, cau_col, total_col


def write_df(ws, df_to_write: pd.DataFrame, freeze_panes: str = "A2") -> None:
    # append dataframe
    for r_idx, row in enumerate(dataframe_to_rows(df_to_write, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:
            for c_idx in range(1, len(row) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = freeze_panes
    ws.auto_filter.ref = ws.dimensions

    # column widths
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max((len(str(c.value)) for c in ws[col_letter] if c.value is not None), default=12)
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 55)

    # auto percent formatting for headers that start with "%_"
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for c_idx, h in enumerate(headers, start=1):
        if isinstance(h, str) and h.strip().startswith("%_"):
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=c_idx).number_format = "0.00%"


def add_legend_box(ws) -> None:
    legend_title = "DIDASCALIA – significato colonne"
    legend_lines = [
        ("Ore_Totali_Anno", "Somma di tutte le ore registrate nell’anno per la risorsa (tutte le causali)."),
        ("Ore_Lavorabili", "Ore potenzialmente lavorabili: Ore_Totali_Anno − Ore_CIG − Ore_Festivita."),
        ("%_Presenza_su_ore_lavorabili", "1 − (Ore_Perse_Persona / Ore_Lavorabili)."),
        ("%_Assenza_su_ore_lavorabili", "Ore_Perse_Persona / Ore_Lavorabili."),
        ("%_CIG_su_totale", "Ore_CIG / Ore_Totali_Anno (solo indicatore)."),
        ("%_Festivita_su_totale", "Ore_Festivita / Ore_Totali_Anno (solo indicatore)."),
        ("Ore_Assenze_Varie", "Somma ore assenze imputabili (malattia, permessi, donazione sangue, assemblea sindacale, ecc.)."),
        ("Ore_Infortunio", "Somma ore di infortunio (tracciato separatamente)."),
        ("Ore_Ferie", "Somma ore di ferie."),
        ("Ore_Perse_Persona", "Ore perse imputabili: Ore_Assenze_Varie + Ore_Infortunio + Ore_Ferie."),
    ]

    start_col = ws.max_column + 2
    start_row = 1

    # merged title across 2 columns
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 1)
    title_cell = ws.cell(row=start_row, column=start_col)
    title_cell.value = legend_title
    title_cell.font = Font(bold=True)
    title_cell.alignment = Alignment(wrap_text=True, vertical="center")

    # header
    ws.cell(row=start_row + 1, column=start_col).value = "Colonna"
    ws.cell(row=start_row + 1, column=start_col + 1).value = "Descrizione"
    ws.cell(row=start_row + 1, column=start_col).font = Font(bold=True)
    ws.cell(row=start_row + 1, column=start_col + 1).font = Font(bold=True)

    r = start_row + 2
    for name, desc in legend_lines:
        ws.cell(row=r, column=start_col).value = name
        ws.cell(row=r, column=start_col + 1).value = desc
        ws.cell(row=r, column=start_col).alignment = Alignment(vertical="top")
        ws.cell(row=r, column=start_col + 1).alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    ws.column_dimensions[get_column_letter(start_col)].width = 30
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 72

    # thin border box
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    end_row = start_row + 1 + len(legend_lines)

    for rr in range(start_row, end_row + 1):
        for cc in range(start_col, start_col + 2):
            ws.cell(row=rr, column=cc).border = border


def run(in_path: Path, out_path: Path) -> None:
    raw = pd.read_excel(in_path, sheet_name=0, header=None)

    header_row = find_header_row(raw)
    ris_col, cau_col, total_col = find_cols(raw, header_row)
    data_start = header_row + 1

    df = raw.iloc[data_start:, :].copy()
    df = df[~(df[cau_col].isna() & df[total_col].isna())].copy()
    df[ris_col] = df[ris_col].ffill()

    df = df.rename(columns={ris_col: "Risorsa", cau_col: "Tipologia", total_col: "Total"})
    df["Risorsa"] = df["Risorsa"].astype(str).str.strip()
    df["Tipologia"] = df["Tipologia"].astype(str).str.strip()
    df["Ore"] = pd.to_numeric(df["Total"], errors="coerce")

    df = df[(df["Risorsa"] != "") & (df["Tipologia"] != "")].copy()
    df = df.dropna(subset=["Ore"])

    df = df[~df["Risorsa"].str.upper().isin(BAD_ROWS)]
    df = df[~df["Tipologia"].str.upper().isin(BAD_ROWS)]

    dettaglio = (
        df.groupby(["Risorsa", "Tipologia"], as_index=False)["Ore"]
          .sum()
          .sort_values(["Risorsa", "Tipologia"])
    )

    pivot = (
        dettaglio.pivot_table(index="Risorsa", columns="Tipologia", values="Ore", aggfunc="sum", fill_value=0)
                .sort_index()
    )
    pivot["TOTALE_ANNO"] = pivot.sum(axis=1)

    # classify columns
    pivot_cols_upper = {c: str(c).strip().upper() for c in pivot.columns}

    unmapped = []
    col_category = {}
    for c in pivot.columns:
        cu = pivot_cols_upper[c]
        if cu == "TOTALE_ANNO":
            continue

        cat = categorize_with_rules(cu)
        if cat is None:
            cat = CAUSALE_CATEGORY.get(cu, None)
        if cat is None:
            unmapped.append(cu)
            cat = "presenza"
        if cat not in VALID_CATS:
            cat = "presenza"
        col_category[c] = cat

    if unmapped:
        print("\nATTENZIONE: causali NON mappate e non catturate dalle regole (trattate come 'presenza'):")
        for u in sorted(set(unmapped)):
            print(" -", u)

    cig_cols = [c for c, cat in col_category.items() if cat == "cig"]
    fest_cols = [c for c, cat in col_category.items() if cat == "festivita"]
    ferie_cols = [c for c, cat in col_category.items() if cat == "ferie"]
    assenze_cols = [c for c, cat in col_category.items() if cat == "assenza_varie"]
    infortunio_cols = [c for c, cat in col_category.items() if cat == "infortunio"]

    ore_tot = pivot["TOTALE_ANNO"].copy()
    ore_cig = pivot[cig_cols].sum(axis=1) if cig_cols else 0
    ore_fest = pivot[fest_cols].sum(axis=1) if fest_cols else 0
    ore_ferie = pivot[ferie_cols].sum(axis=1) if ferie_cols else 0
    ore_assenze = pivot[assenze_cols].sum(axis=1) if assenze_cols else 0
    ore_infortunio = pivot[infortunio_cols].sum(axis=1) if infortunio_cols else 0

    ore_perse_persona = ore_assenze + ore_infortunio + ore_ferie
    ore_lavorabili = ore_tot - ore_cig - ore_fest

    den_tot = ore_tot.replace({0: pd.NA})
    den_lav = ore_lavorabili.replace({0: pd.NA})

    # percentuali (ordine colonne richiesto)
    perc_df = pd.DataFrame({
        "Risorsa": pivot.index,
        "Ore_Totali_Anno": ore_tot.values,
        "Ore_Lavorabili": ore_lavorabili.values if hasattr(ore_lavorabili, "values") else ore_lavorabili,
        "%_Presenza_su_ore_lavorabili": (1 - (ore_perse_persona / den_lav)).astype("Float64"),
        "%_Assenza_su_ore_lavorabili": (ore_perse_persona / den_lav).astype("Float64"),
        "%_CIG_su_totale": (ore_cig / den_tot).astype("Float64"),
        "%_Festivita_su_totale": (ore_fest / den_tot).astype("Float64"),
        "Ore_Assenze_Varie": ore_assenze.values if hasattr(ore_assenze, "values") else ore_assenze,
        "Ore_Infortunio": ore_infortunio.values if hasattr(ore_infortunio, "values") else ore_infortunio,
        "Ore_Ferie": ore_ferie.values if hasattr(ore_ferie, "values") else ore_ferie,
        "Ore_Perse_Persona": ore_perse_persona.values if hasattr(ore_perse_persona, "values") else ore_perse_persona,
    }).reset_index(drop=True)

    # Liste
    risorse = sorted(dettaglio["Risorsa"].unique().tolist())
    tipi = sorted(dettaglio["Tipologia"].unique().tolist())
    max_len = max(len(risorse), len(tipi))
    liste_df = pd.DataFrame({
        "Risorse": (risorse + [""] * (max_len - len(risorse))),
        "": [""] * max_len,
        "Tipologie": (tipi + [""] * (max_len - len(tipi))),
    })

    # Write Excel
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "percentuali presenza"
    write_df(ws1, perc_df, freeze_panes="B2")
    add_legend_box(ws1)

    ws2 = wb.create_sheet("Ore_Annue_Pivot")
    write_df(ws2, pivot.reset_index(), freeze_panes="B2")

    ws3 = wb.create_sheet("Dettaglio_Annuale")
    write_df(ws3, dettaglio, freeze_panes="A2")

    ws4 = wb.create_sheet("Liste")
    write_df(ws4, liste_df, freeze_panes="A2")

    wb.active = 0
    wb.save(out_path)

    print("\nCreato:", out_path)
    print("CIG cols:", cig_cols)
    print("Festività cols:", fest_cols)
    print("Assenze cols:", assenze_cols)
    print("Infortunio cols:", infortunio_cols)
    print("Ferie cols:", ferie_cols)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Genera report presenze annuali da Presenze_Base.xlsx")
    p.add_argument("-i", "--input", required=True, help="Percorso file Presenze_Base.xlsx")
    p.add_argument("-o", "--output", default="Ore_Annue_per_Risorsa_e_Tipologia.xlsx", help="Percorso file output .xlsx")
    return p.parse_args()


if __name__ == "__main__":
    args = parse_args()
    run(Path(args.input), Path(args.output))
